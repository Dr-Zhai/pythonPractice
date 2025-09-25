import openpyxl

excel = openpyxl.load_workbook("../test.xlsx")
sheet = excel["Sheet1"]
rows = sheet.max_row
cols = sheet.max_column
for i in range(1, rows + 1):
    for j in range(1, cols + 1):
        item = sheet.cell(i, j).value
        print("%s" % item, end=" ")
    print()
excel.close()
